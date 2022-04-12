import os
import glob
from openpyxl import *
from pandas import *
from openpyxl.styles import Color, PatternFill, Font, Border

if os.name == "nt":
    sep = "\\"
else:
    sep = "/"

path = os.getcwd()
input_path_1 = path + "/OUTPUT/"
input_path_2 = path + "/DatavyuToSupercoder/Output/"

output_file = glob.glob(os.path.join(input_path_1, "*.xlsx"))[0].split(sep)[-1]
trials_file = glob.glob(os.path.join(input_path_2, "*.xls"))[0].split(sep)[-1]

os.chdir(input_path_1)
output_wb = load_workbook(output_file)
output_sheet = list(list(output_wb)[2])

os.chdir(input_path_2)
trials_wb = read_excel(trials_file, skiprows=[0])
trial_times = trials_wb["Start Time (in elapsed time - Datavyu coding)"]

bad_trials = set()
redFill = PatternFill(start_color = 'FF0000', end_color = 'FF0000', fill_type = 'solid')
for i in range(1, 76):
    row = output_sheet[i]
    for j in range(0, 15):
        cell = row[j]
        if cell.fill == redFill:
            bad_trials.add(i)

txt = "Bad trials:\n"
for trial in sorted(bad_trials):
    time = trial_times[trial-1]
    txt += str(trial) + " (" + time + "), "
    
print(txt[:-2])
