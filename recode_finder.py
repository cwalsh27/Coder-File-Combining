import os
import glob
from openpyxl import load_workbook
from pandas import *
from openpyxl.styles import Color, PatternFill, Font, Border

if os.name == "nt":
    sep = "\\"
else:
    sep = "/"

# gets the study information
with open('config.txt') as f:
    lines = f.readlines()
try:
    study = lines[1][:-1]
except:
    print("No study given in config file. Make sure to put the study type on the second line")
    exit(1)
if not study.strip().lower() in ["facetalk", "wls", "awl", "ewl"]:
    print("Invalid study \"" + study + "\" in config file.")
    exit(1)
else:
    study = study.strip().lower()
try:
    num_trials = int(lines[3].strip())
except:
    print("No number of trials in config file. Make sure to put the number of trials on the fourth line")
    exit(1)
if study in ["facetalk", "wls"]:
    cols = 15
elif study in ["awl"]:
    cols = 21
elif study in ["ewl"]:
    cols = 23

path = os.getcwd()
input_path_1 = path + "/OUTPUT/"
input_path_2 = path + "/DatavyuToSupercoder/Output/"

output_file = glob.glob(os.path.join(input_path_1, "*.xlsx"))[0].split(sep)[-1]
trials_file = glob.glob(os.path.join(input_path_2, "*.xls"))[0].split(sep)[-1]

os.chdir(input_path_1)
output_wb = load_workbook(output_file)
output_sheet = list(list(output_wb)[2])

os.chdir(input_path_2)
trials_wb = read_excel(trials_file, skiprows=[0])
trial_times = trials_wb["Start Time (in elapsed time - Datavyu coding)"]




bad_trials = set()
redFill = PatternFill(start_color = 'FF0000', end_color = 'FF0000', fill_type = 'solid')
redFill2 = PatternFill(start_color = 'FFFF0000', end_color = 'FFFF0000', fill_type = 'solid')
# find the trials marked red
for i in range(1, num_trials+1):
    row = output_sheet[i]
    for j in range(0, cols):
        cell = row[j]
        if cell.fill == redFill or cell.fill == redFill2:
            bad_trials.add(i)

print("\n" + str(len(bad_trials)) + " possible recodes found")

# gets the lines that make up a trial
def trial_region(trial, wb) -> str:
    region1 = ""
    region2 = ""
    first = True
    # get individual regions
    for coder in (0, 1):
        region = ""
        row_num = 0
        sheet = list(list(wb)[coder])
        # find B
        for row in sheet:
            row_num += 1
            cell = row[17].value
            if row[17].value == trial:
                b_row = row_num - 1
                break
        # get everything between B and S
        for i in range(b_row, len(sheet)):               #second error
            cell = sheet[i][0].value
            region += str(i+1) + " "
            if i+1 < 100:
                region += " "
            if i+1 < 10:
                region += " "
            for j in range(0, 3):
                curr_cell = sheet[i][j].value
                if curr_cell:
                    region += str(curr_cell) + " "
                    if j in (1, 2):
                        if curr_cell < 10000:
                            region += " "
                        if curr_cell < 1000:
                            region += " "
                    elif j == 0:
                        if len(curr_cell) == 1:
                            region += " "
                else:
                    region += "      "
            region += "\n"
            if cell == "S":
                break
        if first:
            region1 = region.split("\n")[:-1]
        else:
            region2 = region.split("\n")[:-1]
        first = False
    # combine regions
    length = max(len(region1), len(region2))
    txt = list(wb)[0].title + "            " + list(wb)[1].title + "\n"
    for i in range(length):
        try:
            r1 = region1[i]
        except:
            r1 = "                   "
        try:
            r2 = region2[i]
        except:
            r2 = ""
        txt += r1 + "   " + r2 + "\n"
    return txt
# asks for user input to decide if each trial should be added to the list of recodes
recodes = set()
for trial in sorted(bad_trials):
    print("\nDisagreement between coders detected in trial " + str(trial) + ":")
    print(trial_region(trial, output_wb)[:-1])                                     #first error
    decision = None
    while(not decision in ("1", "2", "3", "4")):
        decision = input("""
If you would like to add this trial to the list of recodes, type 1
If you would like to continue without adding this trial to the list of recodes, type 2
If you would like to stop and manually edit the data, type 3
If you would like to add all trials to the list of recodes, type 4 """)
        
    if decision == "1":
        recodes.add(trial)
    elif decision == "2":
        continue
    elif decision == "3":
        print("\nProcess stopped")
        exit(1)
        break
    else:
        recodes = bad_trials
        break

txt = "Recodes:\n"

    
for trial in sorted(recodes):
    time = trial_times[trial-1]
    txt += str(trial) + " (" + time + "), "
txt = txt[:-2]    

print()
print(txt)

os.chdir(input_path_1)
with open('recodes.txt', 'w') as f:
    f.write(txt)
    
print("Completed! Check OUTPUT folder for list of recodes")
