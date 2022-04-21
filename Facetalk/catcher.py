import os
import glob
from openpyxl import *
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
del open

# "catcher in the py" lol

if os.name == "nt":
    sep = "\\"
else:
    sep = "/"
    
path = os.getcwd() + "/Input/"
config_path = "/".join(os.getcwd().split(sep)[:-1])+"/"
os.chdir(config_path)
with open('config.txt') as f:
    lines = f.readlines()
try:
    study = lines[1][:-1]
except:
    print("No study given in config file. Make sure to put the study type on the second line")
    exit(1)
if not study.lower() in ["facetalk", "wls", "awl"]:
    print("Invalid study \"" + study[:-1] + "\" in config file.")
    exit(1)
try:
    num_trials = int(lines[3])
except:
    print("No number of trials in config file. Make sure to put the number of trials on the fourth line")
    exit(1)

os.chdir

file = glob.glob(os.path.join(path, "*xlsx"))
file_name = file[0].split(sep)[-1]

os.chdir(path)

wb = load_workbook(file_name)

def recreate_line(i, look, on, off) -> str:
    error_row = str(i+1) + " "
    if i+1 < 10:
        error_row += " "
    if i+1 < 100:
        error_row += " "
    error_row += str(look) + "  "
    error_row += str(on) + " "
    if on < 10000:
        error_row +=  " "
    if on < 1000:
        error_row +=  " "
    error_row += str(off)
    return error_row

def error_region(i, sheet, row_num) -> str:
    error_region = "\n"
    # handles edge cases (literally)
    if i == 0:
        area = [i, i+1]
    elif i == row_num - 1:
        area = [i-1, i]
    else:
        area = [i-1, i, i+1]
    # gets the relevant row and two surrounding rows
    for j in area:
        row = list(list(sheet)[j])
        error_region += str(j+1) + " "
        if j+1 < 100:
            error_region += " "
        if j+1 < 10:
            error_region += " "
        for cell in row[0:3]:
            if not cell.value:
                if cell == row[0]:
                    error_region += "  "
                else:
                    error_region += "      "
            else:
                if cell == row[1] or cell == row[2]:
                    if cell.value < 10000:
                        error_region += " "
                    if cell.value < 1000:
                        error_region += " "
                error_region += str(cell.value) + " "
        error_region += "\n"
    return error_region

# catches errors, highlights B cells, and indexes trials
error = False
for sheet in wb:
    b = 0
    s = 0
    row_num = 0
    b_fill = PatternFill(start_color = 'FFD3AA', end_color = 'FFD3AA', fill_type = 'solid')
    if sheet.title == 'AVERAGES ACROSS CODERS':
        wb.remove(sheet)
        wb.save(path + file_name)
        continue
    print("Analyzing " + sheet.title)
    # quick preiminary scan
    for row in sheet:
        if row[0].value == "B":
            b += 1
            # highlights cells and indexes trials
            row[0].fill = b_fill
            sheet["M" + str(row_num + 1)] = b
            sheet["M" + str(row_num + 1)].fill = b_fill
        if row[0].value == "S":
            s += 1
        row_num += 1
        
    # catches errors
    for i in range(0, row_num):
        row = list(sheet)[i]
        # checks offsets and onsets
        if list(row)[0].value in ["B", "S"]:
            if not list(row)[1].value:
                print("\n" + sheet.title + " missing onset in row " + str(i+1) + "\n")
                print(error_region(i, sheet, row_num))
                error = True
            if list(row)[2].value:
                print("\n" + sheet.title + " has offset in row " + str(i+1) + "\n")
                print(error_region(i, sheet, row_num))
                # delete offset for B or S
                print("Suggested fix:")
                print(recreate_line(i, list(row)[0].value, list(row)[1].value, ""))
                approval = input("\nApprove fix? (y/n)")
                if approval.lower() == "y":
                    list(row)[2].value = ""
                else:
                    error = True
        elif list(row)[0].value in ["R", "L", "C", "RT", "RB", "LT", "LB"]:
            if not list(row)[1].value:
                print("\n" + sheet.title + " missing onset in row " + str(i+1) + "\n")
                print(error_region(i, sheet, row_num))
                error = True
            if not list(row)[2].value:
                print("\n" + sheet.title + " missing offset in row " + str(i+1) + "\n")
                print(error_region(i, sheet, row_num))
                error = True
        else:
            print("\n" + sheet.title + " has unrecognized look in row " + str(i+1) + "\n")
            print(error_region(i, sheet, row_num))
            error = True
        # makes sure every S has a B following it (except for the last one)
        if list(row)[0].value == "S" and i != row_num - 1:
            if list(sheet)[i+1][0].value != "B":
                print("\n" + sheet.title + " has an S that isn't followed by a B in row " + str(i+1) + "\n")
                print(error_region(i+1, sheet, row_num))
                error = True
        if list(row)[0].value == "B" and i != 0:
            if list(sheet)[i-1][0].value != "S":
                print("\n" + sheet.title + " has an B that isn't preceded by a S in row " + str(i+1) + "\n")
                print(error_region(i-1, sheet, row_num))
                error = True
    # makes sure the number of trials is correct
    if b != num_trials or s != num_trials:
        print("\n" + sheet.title + " has incorrect number of trials.\n" + str(b) + " B\n" + str(s) + " S\n" + "Should have " + str(num_trials) + " of each." + "\n")       
        error = True
    wb.save(path + file_name)
if error:
    exit(1)